import pandas as pd 

df = pd.DataFrame({
    '상품명': ['Mouse', 'Keyboard', 'Monitor'],
    '수량': [10, 5, 2],
    '가격': [12000, 25000, 300000]
})

# 총금액 컬럼 추가 (수량 * 가격)
df['총금액'] = df['수량'] * df['가격']

#  Excel 파일 생성 : openpyxl 엔진 사용 (엑셀 스타일링 가능)
with pd.ExcelWriter('result5.xlsx', engine='openpyxl') as writer:
    # DataFrame을 엑셀로 저장 (3행부터 시작)
    df.to_excel(writer, sheet_name='Report', index=False, startrow=2)

    # 워크시트 객체 가져오기
    ws = writer.sheets['Report']

    # 제목 추가
    ws['A1'] = '상품 판매 보고서'  # A1 셀에 제목 입력

    from openpyxl.styles import Font  # 글꼴 스타일 모듈

    # 제목 스타일 적용 (굵게 + 글자 크기)
    ws['A1'].font = Font(bold=True, size=14)

    # 헤더 스타일 설정-----
    from openpyxl.styles import PatternFill, Alignment

    # 헤더 글자 스타일 (흰색 + 굵게)
    header_font = Font(bold=True, color='FFFFFF')

    # 헤더 배경색 (파란색)
    header_fill = PatternFill(start_color='4F81BD', fill_type='solid')

    # 3행이 헤더 (startrow=2 → 헤더는 3번째 줄)
    for cell in ws[3]:
        cell.font = header_font              # 글자 스타일 적용
        cell.fill = header_fill                  # 배경색 적용
        cell.alignment = Alignment(horizontal='center')  # 가운데 정렬

    # 컬럼 너비 자동 조정
    for col in ws.columns:  # 모든 컬럼 반복
        max_length = 0
        col_letter = col[0].column_letter  # 컬럼 문자 (A, B, C...)

        for cell in col:
            try:
                if cell.value:
                    # 셀 값 길이 중 최대값 찾기
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # 최대 길이에 여유값 추가하여 너비 설정
        ws.column_dimensions[col_letter].width = max_length + 2

    # 숫자 포맷 적용 (콤마)
    # 데이터 영역: 4행부터 (헤더 다음)
    for row in ws.iter_rows(min_row=4, min_col=2, max_col=4):
        for cell in row:
            # 숫자 타입이면 콤마 포맷 적용
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    # 엑셀 테이블 스타일 추가
    from openpyxl.worksheet.table import Table, TableStyleInfo

    # 테이블 범위 설정 (A3부터 마지막 데이터까지)
    tab = Table(displayName="Table1", ref=f"A3:D{len(df)+3}")

    # 테이블 스타일 지정
    style = TableStyleInfo(
        name="TableStyleMedium9",   # 엑셀 기본 스타일
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,        # 줄무늬
        showColumnStripes=False
    )

    tab.tableStyleInfo = style  # 스타일 적용
    ws.add_table(tab)           # 시트에 테이블 추가

    # 합계 행 추가
    total_row = len(df) + 4  # 합계 위치 계산

    ws[f'A{total_row}'] = '합계'  # 합계 라벨

    # 총금액 합계 (엑셀 함수 사용)
    ws[f'D{total_row}'] = f'=SUM(D4:D{len(df)+3})'

    # 전체 가운데 정렬 (옵션)-
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')